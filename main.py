import logging
from itertools import groupby

import openpyxl
from natsort import natsorted
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill

from read_output_folders import read_output
from result_holder import ResultHolder

logging.basicConfig(level=logging.DEBUG)

logger = logging.getLogger()


def prepare_key(algorithm_aliases, result_holder):
    if "WPM" in result_holder.algorithm and "LWR" not in result_holder.algorithm and "CWR" not in result_holder.algorithm:
        if result_holder.wpm_version == "v1":
            return algorithm_aliases[
                result_holder.path_finding_method + "_" + result_holder.algorithm + "_" + result_holder.wpm_version]
        elif result_holder.wpm_version == "v2":
            return algorithm_aliases[
                result_holder.path_finding_method + "_" + result_holder.algorithm + "_" + result_holder.wpm_version + "_" + result_holder.wpm_value_type]
    elif "WPM" in result_holder.algorithm and "LWR" in result_holder.algorithm and "CWR" not in result_holder.algorithm:
        if result_holder.wpm_version == "v1":
            return algorithm_aliases[
                result_holder.path_finding_method + "_" + result_holder.algorithm + "_" + result_holder.wpm_version]
        elif result_holder.wpm_version == "v2":
            return algorithm_aliases[
                result_holder.path_finding_method + "_" + result_holder.algorithm + "_" + result_holder.wpm_version + "_" + result_holder.wpm_value_type]
    elif "WPM" in result_holder.algorithm and "LWR" in result_holder.algorithm and "CWR" in result_holder.algorithm:
        if result_holder.wpm_version == "v1":
            return algorithm_aliases[
                result_holder.path_finding_method + "_" + result_holder.algorithm + "_" + result_holder.wpm_version]
        elif result_holder.wpm_version == "v2":
            return algorithm_aliases[
                result_holder.path_finding_method + "_" + result_holder.algorithm + "_" + result_holder.wpm_version + "_" + result_holder.wpm_value_type]
    elif "WPM" in result_holder.algorithm and "LWR" not in result_holder.algorithm and "CWR" in result_holder.algorithm:
        if result_holder.wpm_version == "v1":
            return algorithm_aliases[
                result_holder.path_finding_method + "_" + result_holder.algorithm + "_" + result_holder.wpm_version]
        elif result_holder.wpm_version == "v2":
            return algorithm_aliases[
                result_holder.path_finding_method + "_" + result_holder.algorithm + "_" + result_holder.wpm_version + "_" + result_holder.wpm_value_type]
    elif result_holder.path_finding_method == "shortestPath":
        return algorithm_aliases[result_holder.path_finding_method + "_" + result_holder.algorithm]
    elif "U" in result_holder.algorithm:
        return algorithm_aliases[result_holder.path_finding_method + "_" + result_holder.algorithm]
    else:
        return algorithm_aliases[result_holder.path_finding_method + "_" + result_holder.algorithm]


def prepare_value(value):
    value_list = list()
    value_list.append(value.topology_result.topology_name)
    value_list.append(value.topology_result.o1)
    value_list.append(value.topology_result.o2)
    value_list.append(value.topology_result.o3)
    value_list.append(value.topology_result.average_wcd)
    value_list.append(value.topology_result.max_util)
    value_list.append(value.topology_result.average_util)
    value_list.append(value.topology_result.util_variance)
    value_list.append(value.topology_result.time)
    return value_list


def group_output(algorithm_aliases, output_list):
    grouped = groupby(output_list, key=lambda x: (
        x.routing, x.mtr_name, x.path_finding_method, x.algorithm, x.lwr, x.k, x.mcdm_objective, x.wsm_normalization,
        x.cwr, x.w_srt, x.w_tt, x.w_length, x.w_util, x.wpm_version, x.wpm_value_type, x.metaheuristic_name,
        x.evaluator_name))

    grouped_result_dict = {}

    for key, group in grouped:
        k_result_list = {}
        if logger.isEnabledFor(logging.DEBUG):
            logging.debug(key)

        group_sorted = natsorted(list(group), key=lambda x: x.topology_result.topology_name)

        if logger.isEnabledFor(logging.DEBUG):
            for each_group in group_sorted:
                logging.debug(each_group)

        result_holder = ResultHolder()
        result_holder.routing = key[0]
        result_holder.mtr_name = key[1]
        result_holder.path_finding_method = key[2]
        result_holder.algorithm = key[3]
        result_holder.lwr = key[4]
        result_holder.k = key[5]
        result_holder.mcdm_objective = key[6]
        result_holder.wsm_normalization = key[7]
        result_holder.cwr = key[8]
        result_holder.w_srt = key[9]
        result_holder.w_tt = key[10]
        result_holder.w_length = key[11]
        result_holder.w_util = key[12]
        result_holder.wpm_version = key[13]
        result_holder.wpm_value_type = key[14]
        result_holder.metaheuristic_name = key[15]
        result_holder.evaluator_name = key[16]

        new_key = prepare_key(algorithm_aliases, result_holder)

        for each_group in group_sorted:
            if new_key not in k_result_list:
                k_result_list[new_key] = list()
                k_result_list[new_key].append(prepare_value(each_group))
            else:
                k_result_list[new_key].append(prepare_value(each_group))

        if result_holder.k not in grouped_result_dict:
            grouped_result_dict[result_holder.k] = list()
            grouped_result_dict[result_holder.k].append(k_result_list)
        else:
            grouped_result_dict[result_holder.k].append(k_result_list)

    return grouped_result_dict


def sort_grouped_output_list(grouped_output_list):
    first_object = grouped_output_list.pop(None)

    sorted_dict = natsorted(grouped_output_list.items(), key=lambda x: x[0])

    new_dict = dict()
    new_dict[None] = first_object

    for key, value in sorted_dict:
        new_dict[key] = value

    return new_dict


def write_to_excel(wb, algorithm_alias_dict, sheet_list, output_list):
    grouped_output_list = group_output(algorithm_alias_dict, output_list)
    sorted_grouped_output_list = sort_grouped_output_list(grouped_output_list)

    for key, value in sorted_grouped_output_list.items():
        sorted_to_algorithm_aliases = sorted(value,
                                             key=lambda d: list(algorithm_alias_dict.values()).index(list(d.keys())[0]))
        sorted_grouped_output_list[key] = sorted_to_algorithm_aliases

    first_key = list(sorted_grouped_output_list.keys())[1]
    is_first_sheet = True
    for key, value in sorted_grouped_output_list.items():
        for sheet in sheet_list:
            if sheet == sheet_list[0]:
                if is_first_sheet:
                    new_sheet = wb.active
                    new_sheet.title = sheet + "=" + first_key
                    is_first_sheet = False
                else:
                    if sheet + "=" + key in wb.sheetnames:
                        new_sheet = wb[sheet + "=" + key]
                    else:
                        new_sheet = wb.create_sheet(title=sheet + "=" + key)

                max_column = new_sheet.max_column

                col = 0
                if max_column == 1:
                    col = 1
                else:
                    col += max_column + 1

                for output_dict in value:
                    for inner_key, inner_value in output_dict.items():
                        row = 1
                        new_sheet.column_dimensions[
                            new_sheet.cell(row=row, column=col + 1, value=inner_key).column_letter].width = len(
                            inner_key) + 5
                        row += 1
                        for result in inner_value:
                            column_number = 0
                            new_sheet.cell(row=row, column=col + column_number, value=result[0])
                            column_number += 1
                            new_sheet.cell(row=row, column=col + column_number, value=float(result[1]))
                            column_number += 1
                            row += 1

                        col += len(inner_value[0][:2])

            elif sheet == sheet_list[1]:
                if key is not None:
                    if sheet + "=" + key in wb.sheetnames:
                        new_sheet = wb[sheet + "=" + key]
                    else:
                        new_sheet = wb.create_sheet(title=sheet + "=" + key)
                else:
                    new_sheet = wb.create_sheet(title=sheet + "=" + first_key)

                max_column = new_sheet.max_column

                col = 0
                if max_column == 1:
                    col = 1
                else:
                    col += max_column + 1

                for output_dict in value:
                    for inner_key, inner_value in output_dict.items():
                        row = 1
                        new_sheet.column_dimensions[
                            new_sheet.cell(row=row, column=col + 1, value=inner_key).column_letter].width = len(
                            inner_key) + 5
                        row += 1
                        for result in inner_value:
                            column_number = 0
                            new_sheet.cell(row=row, column=col + column_number, value=result[0])
                            column_number += 1
                            new_sheet.cell(row=row, column=col + column_number, value=float(result[2]))
                            column_number += 1
                            new_sheet.cell(row=row, column=col + column_number, value=float(result[3]))
                            column_number += 1
                            row += 1

                        col += len(inner_value[0][2:4]) + 1

            elif sheet == sheet_list[2]:
                if key is not None:
                    if sheet + "=" + key in wb.sheetnames:
                        new_sheet = wb[sheet + "=" + key]
                    else:
                        new_sheet = wb.create_sheet(title=sheet + "=" + key)
                else:
                    new_sheet = wb.create_sheet(title=sheet + "=" + first_key)

                max_column = new_sheet.max_column

                col = 0
                if max_column == 1:
                    col = 1
                else:
                    col += max_column + 1

                for output_dict in value:
                    for inner_key, inner_value in output_dict.items():
                        row = 1
                        new_sheet.column_dimensions[
                            new_sheet.cell(row=row, column=col + 1, value=inner_key).column_letter].width = len(
                            inner_key) + 5
                        row += 1
                        for result in inner_value:
                            column_number = 0
                            new_sheet.cell(row=row, column=col + column_number, value=result[0])
                            column_number += 1
                            new_sheet.cell(row=row, column=col + column_number, value=float(result[4]))
                            column_number += 1
                            row += 1

                        col += len(inner_value[0][:2])

            elif sheet == sheet_list[3]:
                if key is not None:
                    if sheet + "=" + key in wb.sheetnames:
                        new_sheet = wb[sheet + "=" + key]
                    else:
                        new_sheet = wb.create_sheet(title=sheet + "=" + key)
                else:
                    new_sheet = wb.create_sheet(title=sheet + "=" + first_key)

                max_column = new_sheet.max_column

                col = 0
                if max_column == 1:
                    col = 1
                else:
                    col += max_column + 1

                for output_dict in value:
                    for inner_key, inner_value in output_dict.items():
                        row = 1
                        new_sheet.column_dimensions[
                            new_sheet.cell(row=row, column=col + 1, value=inner_key).column_letter].width = len(
                            inner_key) + 5
                        row += 1
                        for result in inner_value:
                            column_number = 0
                            new_sheet.cell(row=row, column=col + column_number, value=result[0])
                            column_number += 1
                            new_sheet.cell(row=row, column=col + column_number, value=float(result[5]))
                            column_number += 1
                            new_sheet.cell(row=row, column=col + column_number, value=float(result[6]))
                            column_number += 1
                            new_sheet.cell(row=row, column=col + column_number, value=float(result[7]))
                            column_number += 1
                            new_sheet.cell(row=row, column=col + column_number, value=float(result[8]))
                            column_number += 1
                            row += 1

                        col += len(inner_value[0][5:9]) + 1

    wb.save("output.xlsx")


def find_min_value_keys(d):
    # Find the minimum value in the dictionary
    min_value = min(d.values())

    # Find all keys that have the minimum value
    min_keys = [k for k, v in d.items() if v == min_value]

    # Create a dictionary with these keys and their values
    min_value_dict = {k: min_value for k in min_keys}

    return min_value_dict


def paint_cells(green_fill, yellow_fill, red_fill, workbook):
    i = 0
    for sheet in workbook.worksheets:
        if "O1-K" in sheet.title:
            j = 2
            for row in sheet.iter_rows(min_row=2):
                if i < 3:

                    ro_cell = row[5]

                    wpm_pp_cell = row[7]
                    wpm_yen_cell = row[9]
                    lwr_pp_cell = row[11]
                    lwr_yen_cell = row[13]
                    cwr_pp_cell = row[15]
                    cwr_yen_cell = row[17]
                    lwr_cwr_pp_cell = row[19]
                    lwr_cwr_yen_cell = row[21]

                    ro_cell_value = float(ro_cell.value)

                    cell_value_dict = dict()

                    cell_value_dict[wpm_pp_cell] = float(wpm_pp_cell.value)
                    cell_value_dict[wpm_yen_cell] = float(wpm_yen_cell.value)
                    cell_value_dict[lwr_pp_cell] = float(lwr_pp_cell.value)
                    cell_value_dict[lwr_yen_cell] = float(lwr_yen_cell.value)
                    cell_value_dict[cwr_pp_cell] = float(cwr_pp_cell.value)
                    cell_value_dict[cwr_yen_cell] = float(cwr_yen_cell.value)
                    cell_value_dict[lwr_cwr_pp_cell] = float(lwr_cwr_pp_cell.value)
                    cell_value_dict[lwr_cwr_yen_cell] = float(lwr_cwr_yen_cell.value)

                    min_value_keys_dict = find_min_value_keys(cell_value_dict)

                    for key, value in min_value_keys_dict.items():
                        if value < ro_cell_value:
                            key.fill = green_fill
                        elif value == ro_cell_value:
                            key.fill = yellow_fill
                        else:
                            key.fill = red_fill

                    ip_col = 2
                    if ro_cell_value != 0:
                        sheet.cell(row=1, column=21 + ip_col, value="ip_lwr_pp")
                        ip_lwr_pp = ((ro_cell_value - lwr_pp_cell.value) * 100) / ro_cell_value
                        sheet.cell(row=j, column=21 + ip_col, value=float(ip_lwr_pp))

                        ip_col += 1

                        sheet.cell(row=1, column=21 + ip_col, value="ip_lwr_yen")
                        ip_lwr_yen = ((ro_cell_value - lwr_yen_cell.value) * 100) / ro_cell_value
                        sheet.cell(row=j, column=21 + ip_col, value=float(ip_lwr_yen))

                        ip_col += 1

                        sheet.cell(row=1, column=21 + ip_col, value="ip_cwr_pp")
                        ip_cwr_pp = ((ro_cell_value - cwr_pp_cell.value) * 100) / ro_cell_value
                        sheet.cell(row=j, column=21 + ip_col, value=float(ip_cwr_pp))

                        ip_col += 1

                        sheet.cell(row=1, column=21 + ip_col, value="ip_cwr_yen")
                        ip_cwr_yen = ((ro_cell_value - cwr_yen_cell.value) * 100) / ro_cell_value
                        sheet.cell(row=j, column=21 + ip_col, value=float(ip_cwr_yen))

                        ip_col += 1

                        sheet.cell(row=1, column=21 + ip_col, value="ip_lwr_cwr_pp")
                        ip_lwr_cwr_pp = ((ro_cell_value - lwr_cwr_pp_cell.value) * 100) / ro_cell_value
                        sheet.cell(row=j, column=21 + ip_col, value=float(ip_lwr_cwr_pp))

                        ip_col += 1

                        sheet.cell(row=1, column=21 + ip_col, value="ip_lwr_cwr_yen")
                        ip_lwr_cwr_yen = ((ro_cell_value - lwr_cwr_yen_cell.value) * 100) / ro_cell_value
                        sheet.cell(row=j, column=21 + ip_col, value=float(ip_lwr_cwr_yen))

                    j += 1

                else:

                    ro_cell = row[3]

                    wpm_pp_cell = row[5]
                    wpm_yen_cell = row[7]
                    lwr_pp_cell = row[9]
                    lwr_yen_cell = row[11]
                    cwr_pp_cell = row[13]
                    cwr_yen_cell = row[15]
                    lwr_cwr_pp_cell = row[17]
                    lwr_cwr_yen_cell = row[19]

                    ro_cell_value = float(ro_cell.value)

                    cell_value_dict = dict()

                    cell_value_dict[wpm_pp_cell] = float(wpm_pp_cell.value)
                    cell_value_dict[wpm_yen_cell] = float(wpm_yen_cell.value)
                    cell_value_dict[lwr_pp_cell] = float(lwr_pp_cell.value)
                    cell_value_dict[lwr_yen_cell] = float(lwr_yen_cell.value)
                    cell_value_dict[cwr_pp_cell] = float(cwr_pp_cell.value)
                    cell_value_dict[cwr_yen_cell] = float(cwr_yen_cell.value)
                    cell_value_dict[lwr_cwr_pp_cell] = float(lwr_cwr_pp_cell.value)
                    cell_value_dict[lwr_cwr_yen_cell] = float(lwr_cwr_yen_cell.value)

                    min_value_keys_dict = find_min_value_keys(cell_value_dict)

                    for key, value in min_value_keys_dict.items():
                        if value < ro_cell_value:
                            key.fill = green_fill
                        elif value == ro_cell_value:
                            key.fill = yellow_fill
                        else:
                            key.fill = red_fill

                    ip_col = 2
                    if ro_cell_value != 0:
                        sheet.cell(row=1, column=19 + ip_col, value="ip_lwr_pp")
                        ip_lwr_pp = ((ro_cell_value - lwr_pp_cell.value) * 100) / ro_cell_value
                        sheet.cell(row=j, column=19 + ip_col, value=float(ip_lwr_pp))

                        ip_col += 1

                        sheet.cell(row=1, column=19 + ip_col, value="ip_lwr_yen")
                        ip_lwr_yen = ((ro_cell_value - lwr_yen_cell.value) * 100) / ro_cell_value
                        sheet.cell(row=j, column=19 + ip_col, value=float(ip_lwr_yen))

                        ip_col += 1

                        sheet.cell(row=1, column=19 + ip_col, value="ip_cwr_pp")
                        ip_cwr_pp = ((ro_cell_value - cwr_pp_cell.value) * 100) / ro_cell_value
                        sheet.cell(row=j, column=19 + ip_col, value=float(ip_cwr_pp))

                        ip_col += 1

                        sheet.cell(row=1, column=19 + ip_col, value="ip_cwr_yen")
                        ip_cwr_yen = ((ro_cell_value - cwr_yen_cell.value) * 100) / ro_cell_value
                        sheet.cell(row=j, column=19 + ip_col, value=float(ip_cwr_yen))

                        ip_col += 1

                        sheet.cell(row=1, column=19 + ip_col, value="ip_lwr_cwr_pp")
                        ip_lwr_cwr_pp = ((ro_cell_value - lwr_cwr_pp_cell.value) * 100) / ro_cell_value
                        sheet.cell(row=j, column=19 + ip_col, value=float(ip_lwr_cwr_pp))

                        ip_col += 1

                        sheet.cell(row=1, column=19 + ip_col, value="ip_lwr_cwr_yen")
                        ip_lwr_cwr_yen = ((ro_cell_value - lwr_cwr_yen_cell.value) * 100) / ro_cell_value
                        sheet.cell(row=j, column=19 + ip_col, value=float(ip_lwr_cwr_yen))

                    j += 1

            i += 1

        if "O2-O3-K" in sheet.title:
            for row in sheet.iter_rows(min_row=2):
                pass
            i += 1

        if "U-T-K" in sheet.title:
            for row in sheet.iter_rows(min_row=2):
                pass

            i += 1

    workbook.save("output_v2_relative_colored.xlsx")


def main():
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="00FF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    wb = openpyxl.Workbook()

    project = "tsn-simulation"

    algorithm_alias_dict = {"shortestPath_dijkstra": "Dijkstra",
                            "yen_U": "KSPU",

                            "yen_GRASP": "RO",

                            "pathPenalization_WSM": "WSM_pp",
                            "yen_WSM": "WSM_yen",

                            "pathPenalization_WSMLWR": "WSM_LWR_pp",
                            "yen_WSMLWR": "WSM_LWR_yen",

                            "pathPenalization_WSMCWR": "WSM_CWR_pp",
                            "yen_WSMCWR": "WSM_CWR_yen",

                            "pathPenalization_WSMLWRCWR": "WSM_LWR_CWR_pp",
                            "yen_WSMLWRCWR": "WSM_LWR_CWR_yen",

                            "pathPenalization_WPM_v1": "WPM_pp_v1",
                            "pathPenalization_WPM_v2_actual": "WPM_pp_v2_actual",
                            "pathPenalization_WPM_v2_relative": "WPM_pp_v2_relative",
                            "yen_WPM_v1": "WPM_yen_v1",
                            "yen_WPM_v2_actual": "WPM_yen_v2_actual",
                            "yen_WPM_v2_relative": "WPM_yen_v2_relative",

                            "pathPenalization_WPMLWR_v1": "LWR_pp_v1",
                            "pathPenalization_WPMLWR_v2_actual": "LWR_pp_v2_actual",
                            "pathPenalization_WPMLWR_v2_relative": "LWR_pp_v2_relative",
                            "yen_WPMLWR_v1": "LWR_yen_v1",
                            "yen_WPMLWR_v2_actual": "LWR_yen_v2_actual",
                            "yen_WPMLWR_v2_relative": "LWR_yen_v2_relative",

                            "pathPenalization_WPMCWR_v1": "CWR_pp_v1",
                            "pathPenalization_WPMCWR_v2_actual": "CWR_pp_v2_actual",
                            "pathPenalization_WPMCWR_v2_relative": "CWR_pp_v2_relative",
                            "yen_WPMCWR_v1": "CWR_yen_v1",
                            "yen_WPMCWR_v2_actual": "CWR_yen_v2_actual",
                            "yen_WPMCWR_v2_relative": "CWR_yen_v2_relative",

                            "pathPenalization_WPMLWRCWR_v1": "LWR_CWR_pp_v1",
                            "pathPenalization_WPMLWRCWR_v2_actual": "LWR_CWR_pp_v2_actual",
                            "pathPenalization_WPMLWRCWR_v2_relative": "LWR_CWR_pp_v2_relative",
                            "yen_WPMLWRCWR_v1": "LWR_CWR_yen_v1",
                            "yen_WPMLWRCWR_v2_actual": "LWR_CWR_yen_v2_actual",
                            "yen_WPMLWRCWR_v2_relative": "LWR_CWR_yen_v2_relative",

                            }

    sheet_list = ["O1-K", "O2-O3-K", "AWCD-K", "MU-AU-V-T-K"]
    output_list = read_output(project)

    write_to_excel(wb, algorithm_alias_dict, sheet_list, output_list)

    workbook = load_workbook("output_v2_relative.xlsx")

    # paint_cells(green_fill, yellow_fill, red_fill, workbook)


if __name__ == "__main__":
    main()
